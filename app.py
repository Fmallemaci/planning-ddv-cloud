
from __future__ import annotations

import io
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from supabase import Client, create_client

APP_DIR = Path(__file__).resolve().parent
CHESS_COLUMNS = [
    "idCns","dsCns","TotPDV","TotBlt","TotUPs","TotVal","TotFdR",
    "TotUdT","TotDia","TotUdM","TotPes","TotPkg","TotDSB","TotDSV",
    "TotDSU","TotCrg","TotCbt"
]
ROLE_MAP = {"CHOFER":"CHOFER", "AYUDANTE 1":"AYUDANTE_1", "AYUDANTE 2":"AYUDANTE_2"}
DIV_NAMES = {"TW":"TRELEW", "PM":"PUERTO MADRYN"}

st.set_page_config(page_title="Planning DDV", page_icon="🚚", layout="wide", initial_sidebar_state="expanded")

CSS = """
<style>
:root {
  --bg:#071927; --panel:#102b3e; --panel2:#24323c; --line:#35536a;
  --tw:#2c7a24; --pm:#0c67ad; --valle:#58656e; --cyan:#35d2da;
}
html,body,[data-testid="stAppViewContainer"] {background:linear-gradient(135deg,#06131f,#0a2333)!important;color:#f4f8fb;}
[data-testid="stSidebar"] {background:linear-gradient(180deg,#263640,#172833)!important;border-right:1px solid #45616f;}
[data-testid="stSidebar"] * {color:#f3f7fa!important;}
.block-container {max-width:1680px;padding-top:1rem;padding-bottom:2rem;}
h1,h2,h3 {letter-spacing:-.02em;}
.hero {
 min-height:180px;border-radius:16px;overflow:hidden;display:flex;align-items:center;justify-content:center;
 background:
 linear-gradient(90deg,rgba(4,16,26,.1),rgba(4,16,26,.93) 42%,rgba(4,16,26,.25)),
 url('https://images.unsplash.com/photo-1601584115197-04ecc0da31d7?auto=format&fit=crop&w=1800&q=80') center/cover;
 border:1px solid #436170;box-shadow:0 12px 30px rgba(0,0,0,.22);
}
.hero-title{text-align:center;text-shadow:0 2px 12px #000;font-weight:900;font-size:44px}
.hero-sub{font-size:18px;color:#53d8ff;font-weight:800;letter-spacing:.14em}
.card {border:1px solid var(--line);border-radius:14px;padding:18px;background:rgba(14,39,56,.92);box-shadow:0 10px 26px rgba(0,0,0,.18)}
.valle {background:linear-gradient(135deg,#4b5962,#263742);border-color:#7b8991}
.pm {background:linear-gradient(135deg,#084776,#073354);border-color:#1585d2}
.tw {background:linear-gradient(135deg,#1b5f20,#103c18);border-color:#50a94b}
.kpi-grid{display:grid;grid-template-columns:repeat(6,minmax(120px,1fr));gap:1px;background:#5d7482;border:1px solid #758a95;border-radius:12px;overflow:hidden}
.kpi{padding:16px 12px;text-align:center;background:rgba(6,25,39,.93)}
.kpi .ico{font-size:28px}.kpi .v{font-size:31px;font-weight:900}.kpi .l{font-size:12px;font-weight:800;text-transform:uppercase;color:#cfe0e9}
.base-head{padding:12px 16px;border-radius:12px 12px 0 0;font-size:22px;font-weight:900}
.base-head.twc{background:#247021}.base-head.pmc{background:#0b65a8}
.metric-row{display:grid;grid-template-columns:repeat(6,1fr);background:#fff;color:#092238;border-radius:0 0 12px 12px;overflow:hidden}
.metric-box{text-align:center;padding:13px 7px;border-right:1px solid #d3dde4}
.metric-box:last-child{border:none}.metric-box .n{font-size:25px;font-weight:900}.metric-box .t{font-size:10px;font-weight:800;text-transform:uppercase}
.status-ok{color:#42e777;font-weight:800}.status-pending{color:#ffca58;font-weight:800}
div[data-testid="stDataFrame"]{border:1px solid #35566b;border-radius:10px;overflow:hidden}
.stButton>button,.stDownloadButton>button {border-radius:9px;font-weight:800;min-height:42px}
.stButton>button[kind="primary"]{background:#0876ca;border:none}
[data-testid="stMetric"]{background:rgba(255,255,255,.06);padding:12px;border-radius:10px;border:1px solid #39566a}
@media(max-width:1100px){.kpi-grid,.metric-row{grid-template-columns:repeat(3,1fr)}.hero-title{font-size:32px}}
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

def canon(v: Any) -> str:
    return re.sub(r"\s+"," ",str(v or "").strip().upper())

def num(v: Any) -> float:
    if v in (None,""): return 0.0
    try: return float(str(v).replace(",","."))
    except: return 0.0

@st.cache_resource
def sb() -> Client:
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_SECRET_KEY"]
    return create_client(url, key)

def secret_users() -> dict:
    users = {}
    for key in ("fernando","kevin","nestor"):
        if key in st.secrets.get("users", {}):
            raw = st.secrets["users"][key]
            users[key] = {
                "password": str(raw.get("password","")),
                "name": str(raw.get("name",key.title())),
                "role": str(raw.get("role","operador")),
                "division": str(raw.get("division",""))
            }
    return users

def login():
    if st.session_state.get("auth"): return
    st.markdown('<div class="hero"><div><div class="hero-title">PLANNING DDV</div><div class="hero-sub">CONTROL OPERATIVO CHESS</div></div></div>', unsafe_allow_html=True)
    st.write("")
    c1,c2,c3 = st.columns([1,1.2,1])
    with c2:
        with st.container(border=True):
            st.subheader("Ingreso")
            username = st.text_input("Usuario").strip().lower()
            password = st.text_input("Contraseña", type="password")
            if st.button("Ingresar", type="primary", use_container_width=True):
                u = secret_users().get(username)
                if u and password == u["password"]:
                    st.session_state.auth = {"username":username, **u}
                    st.rerun()
                else:
                    st.error("Usuario o contraseña incorrectos.")
    st.stop()

login()
user = st.session_state.auth

with st.sidebar:
    st.markdown("## 🚚 Planning DDV")
    st.caption("Control de distribución")
    st.markdown(f"**{user['name']}**  \n{user['role'].title()}")
    st.divider()
    page = st.radio("Navegación", ["Dashboard","Planning CHESS","Salida diaria","Empleados","Histórico"], label_visibility="collapsed")
    st.divider()
    if st.button("Cerrar sesión", use_container_width=True):
        st.session_state.clear(); st.rerun()

def get_day(day: date, create=False):
    q = sb().table("planning_days").select("*").eq("operational_date", day.isoformat()).execute().data
    if q: return q[0]
    if not create: return None
    return sb().table("planning_days").insert({
        "operational_date": day.isoformat(), "general_status":"borrador"
    }).execute().data[0]

def ensure_div_status(day_id, division):
    q = sb().table("division_status").select("*").eq("planning_day_id",day_id).eq("division",division).execute().data
    if q:return q[0]
    return sb().table("division_status").insert({
        "planning_day_id":day_id,"division":division,"status":"sin_cargar"
    }).execute().data[0]

def routes_for(day: date, division=None):
    d = get_day(day)
    if not d:return []
    q = sb().table("planning_routes").select("*,locations(name)").eq("planning_day_id",d["id"])
    if division in ("TW","PM"): q=q.eq("division",division)
    return q.order("division").order("unit_id").execute().data

def assignments_for(route_ids):
    if not route_ids:return []
    return sb().table("route_assignments").select("*,employees(full_name)").in_("route_id",route_ids).execute().data

def employees():
    rows = sb().table("employees").select("*,employee_divisions(division)").order("full_name").execute().data
    for r in rows:
        r["divisions"]=[x["division"] for x in r.get("employee_divisions",[])]
    return rows

def locations():
    return sb().table("locations").select("*").eq("active",True).order("division").order("display_order").execute().data

def read_chess(file) -> pd.DataFrame:
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: raise ValueError("Archivo vacío.")
    headers = [str(x or "").strip() for x in rows[0]]
    idx = {h:i for i,h in enumerate(headers)}
    missing=[c for c in CHESS_COLUMNS if c not in idx]
    if missing: raise ValueError("Faltan columnas: "+", ".join(missing))
    data=[]
    for row in rows[1:]:
        if not any(v not in (None,"") for v in row): continue
        data.append({c:row[idx[c]] for c in CHESS_COLUMNS})
    return pd.DataFrame(data)

def upsert_file(day: date, division: str, file):
    df=read_chess(file)
    d=get_day(day,True); ensure_div_status(d["id"],division)
    old=routes_for(day,division)
    old_map={(canon(r["unit_id"]),canon(r["domain"])):r for r in old}
    seen=set()
    for i,row in df.iterrows():
        key=(canon(row["idCns"]),canon(row["dsCns"])); seen.add(key)
        payload={
            "planning_day_id":d["id"],"division":division,"source_row":int(i+2),
            "unit_id":str(row["idCns"]).strip(),"domain":canon(row["dsCns"]),
            "pdv":num(row["TotPDV"]),"packages":num(row["TotBlt"]),
            "pure_pallets":num(row["TotUPs"]),"amount_to_collect":num(row["TotVal"]),
            "out_of_route":num(row["TotFdR"]),"total_udt":num(row["TotUdT"]),
            "total_day":num(row["TotDia"]),"hectoliters":num(row["TotUdM"]),
            "weight":num(row["TotPes"]),"picking":num(row["TotPkg"]),
            "avg_dsb":num(row["TotDSB"]),"avg_dsv":num(row["TotDSV"]),
            "avg_dsu":num(row["TotDSU"]),"total_loads":num(row["TotCrg"]),
            "receipts":num(row["TotCbt"]),"updated_at":datetime.utcnow().isoformat()
        }
        if key in old_map:
            sb().table("planning_routes").update(payload).eq("id",old_map[key]["id"]).execute()
        else:
            sb().table("planning_routes").insert(payload).execute()
    # elimina rutas obsoletas de esa división al reemplazar archivo
    for key,r in old_map.items():
        if key not in seen:
            sb().table("planning_routes").delete().eq("id",r["id"]).execute()
    sb().table("division_status").update({"status":"borrador","updated_at":datetime.utcnow().isoformat()}).eq("planning_day_id",d["id"]).eq("division",division).execute()
    sb().table("import_batches").insert({
        "planning_day_id":d["id"],"division":division,
        "original_filename":getattr(file,"name","archivo.xlsx"),"row_count":len(df)
    }).execute()
    return len(df)

def day_metrics(day: date):
    rr=routes_for(day)
    out={}
    for div in ("TW","PM"):
        rows=[r for r in rr if r["division"]==div]
        out[div]={"cam":len(rows),"pdv":sum(num(r["pdv"]) for r in rows),"bultos":sum(num(r["packages"]) for r in rows)}
    out["ALL"]={k:out["TW"][k]+out["PM"][k] for k in ("cam","pdv","bultos")}
    return out,rr

def kpi_html(m):
    ds=(m["bultos"]/m["pdv"]) if m["pdv"] else 0
    return f"""<div class="kpi-grid">
      <div class="kpi"><div class="ico">🚚</div><div class="v">{m['cam']}</div><div class="l">Camiones</div></div>
      <div class="kpi"><div class="ico">🏪</div><div class="v">{m['pdv']:.0f}</div><div class="l">PDV</div></div>
      <div class="kpi"><div class="ico">📦</div><div class="v">{m['bultos']:,.1f}</div><div class="l">Bultos</div></div>
      <div class="kpi"><div class="ico">🎯</div><div class="v">{ds:.1f}</div><div class="l">Drop size</div></div>
      <div class="kpi"><div class="ico">👥</div><div class="v">{max(11-m['cam'],0)}</div><div class="l">Flota sin asignación</div></div>
      <div class="kpi"><div class="ico">📊</div><div class="v">{(m['cam']/11*100):.1f}%</div><div class="l">Utilización total</div></div>
    </div>"""

def base_card(div, metric, rows):
    cap=6 if div=="TW" else 5
    color="twc" if div=="TW" else "pmc"
    util=metric["cam"]/cap*100 if cap else 0
    ds=metric["bultos"]/metric["pdv"] if metric["pdv"] else 0
    st.markdown(f'<div class="base-head {color}">📍 {DIV_NAMES[div]} <span style="float:right;font-size:13px">{metric["cam"]} unidades</span></div>', unsafe_allow_html=True)
    st.markdown(f"""<div class="metric-row">
      <div class="metric-box"><div class="n">🚚 {metric['cam']}</div><div class="t">Camiones</div></div>
      <div class="metric-box"><div class="n">👤 {max(cap-metric['cam'],0)}</div><div class="t">Sin asignación</div></div>
      <div class="metric-box"><div class="n">◔ {util:.1f}%</div><div class="t">Utilización</div></div>
      <div class="metric-box"><div class="n">🏪 {metric['pdv']:.0f}</div><div class="t">PDV</div></div>
      <div class="metric-box"><div class="n">📦 {metric['bultos']:,.1f}</div><div class="t">Bultos</div></div>
      <div class="metric-box"><div class="n">🎯 {ds:.1f}</div><div class="t">Drop size</div></div>
    </div>""", unsafe_allow_html=True)
    if rows:
        df=pd.DataFrame([{
            "UNIDAD":r["unit_id"],"DOMINIO":r["domain"],"PDV":r["pdv"],
            "LOCALIDAD":(r.get("locations") or {}).get("name",""),
            "BULTOS":round(num(r["packages"]),1),"ESTADO":"Completo" if r.get("rendition_number") else "Pendiente"
        } for r in rows])
        st.dataframe(df,hide_index=True,use_container_width=True,height=min(330,40+35*len(df)))

def header():
    st.markdown('<div class="hero"><div><div class="hero-title">DDV »<br>PLANNING DDV</div><div class="hero-sub">CONTROL OPERATIVO CHESS</div></div></div>', unsafe_allow_html=True)

if page=="Dashboard":
    header()
    c1,c2,c3=st.columns([1,1,2])
    with c1: day=st.date_input("Fecha operativa", value=date.today())
    with c2: scope=st.selectbox("División",["TODAS","TW","PM"])
    metrics,rr=day_metrics(day)
    st.markdown('<div class="card valle"><h3>VALLE · DDV TOTAL</h3>'+kpi_html(metrics["ALL"])+'</div>',unsafe_allow_html=True)
    st.write("")
    a,b=st.columns(2)
    with a:
        with st.container(border=True):
            base_card("PM",metrics["PM"],[r for r in rr if r["division"]=="PM"])
    with b:
        with st.container(border=True):
            base_card("TW",metrics["TW"],[r for r in rr if r["division"]=="TW"])

elif page=="Planning CHESS":
    st.title("Planning CHESS")
    a,b,c=st.columns([1,1,2])
    with a: day=st.date_input("Fecha",value=date.today())
    with b:
        allowed=["TW","PM"] if user["role"]=="admin" or not user["division"] else [user["division"]]
        division=st.selectbox("División",allowed,index=0)
    with c:
        file=st.file_uploader(f"Archivo {division}",type=["xlsx"])
        if file and st.button("Importar / actualizar",type="primary"):
            try:
                n=upsert_file(day,division,file); st.success(f"{n} salidas importadas."); st.rerun()
            except Exception as e: st.error(str(e))
    rr=routes_for(day,division)
    if not rr:
        st.info("No hay salidas cargadas para la fecha y división.")
    else:
        emps=employees(); locs=[x for x in locations() if x["division"]==division]
        emp_name={e["id"]:e["full_name"] for e in emps}
        loc_name={l["id"]:l["name"] for l in locs}
        eligible_driver=[""]+[e["full_name"] for e in emps if e["active"] and e["can_drive"] and division in e["divisions"]]
        eligible_helper=[""]+[e["full_name"] for e in emps if e["active"] and e["can_assist"] and division in e["divisions"]]
        ass=assignments_for([r["id"] for r in rr])
        amap={}
        for x in ass:
            amap.setdefault(x["route_id"],{})[x["role"]]=(x.get("employees") or {}).get("full_name","")
        rows=[]
        for r in rr:
            rows.append({
                "_id":r["id"],"DIV":r["division"],"UNIDAD":r["unit_id"],"DOMINIO":r["domain"],
                "PDV":r["pdv"],"BULTOS":r["packages"],"RENDICIÓN":r.get("rendition_number") or "",
                "CHOFER":amap.get(r["id"],{}).get("CHOFER",""),
                "AYUDANTE 1":amap.get(r["id"],{}).get("AYUDANTE_1",""),
                "AYUDANTE 2":amap.get(r["id"],{}).get("AYUDANTE_2",""),
                "LOCALIDAD":loc_name.get(r.get("location_id"),""),
                "OBSERVACIONES":r.get("observations") or ""
            })
        df=pd.DataFrame(rows)
        edited=st.data_editor(
            df,hide_index=True,use_container_width=True,num_rows="fixed",
            disabled=["_id","DIV","UNIDAD","DOMINIO","PDV","BULTOS"],
            column_config={
                "_id":None,
                "CHOFER":st.column_config.SelectboxColumn(options=eligible_driver),
                "AYUDANTE 1":st.column_config.SelectboxColumn(options=eligible_helper),
                "AYUDANTE 2":st.column_config.SelectboxColumn(options=eligible_helper),
                "LOCALIDAD":st.column_config.SelectboxColumn(options=[""]+[l["name"] for l in locs]),
            }
        )
        c1,c2=st.columns([1,1])
        if c1.button("Guardar borrador",type="primary",use_container_width=True):
            try:
                name_to_emp={e["full_name"]:e["id"] for e in emps}
                name_to_loc={l["name"]:l["id"] for l in locs}
                for _,row in edited.iterrows():
                    rid=row["_id"]
                    sb().table("planning_routes").update({
                        "rendition_number":str(row["RENDICIÓN"] or ""),
                        "location_id":name_to_loc.get(row["LOCALIDAD"]) if row["LOCALIDAD"] else None,
                        "observations":str(row["OBSERVACIONES"] or ""),
                        "is_cyo":"CYO" in canon(row["DOMINIO"]) or "CYO" in canon(row["LOCALIDAD"]) or "CYO" in canon(row["OBSERVACIONES"]),
                        "updated_at":datetime.utcnow().isoformat()
                    }).eq("id",rid).execute()
                    # reemplazar asignaciones de la ruta
                    sb().table("route_assignments").delete().eq("route_id",rid).execute()
                    for col,role in ROLE_MAP.items():
                        nm=str(row[col] or "").strip()
                        if nm:
                            sb().table("route_assignments").insert({"route_id":rid,"employee_id":name_to_emp[nm],"role":role}).execute()
                d=get_day(day)
                sb().table("division_status").update({"status":"borrador","updated_at":datetime.utcnow().isoformat()}).eq("planning_day_id",d["id"]).eq("division",division).execute()
                st.success("Borrador guardado en Supabase.")
            except Exception as e: st.error(f"No se pudo guardar: {e}")
        if c2.button("Confirmar división",use_container_width=True):
            d=get_day(day)
            sb().table("division_status").update({"status":"confirmado","confirmed_at":datetime.utcnow().isoformat()}).eq("planning_day_id",d["id"]).eq("division",division).execute()
            st.success(f"{DIV_NAMES[division]} confirmado.")

elif page=="Salida diaria":
    st.title("Salida diaria")
    c1,c2=st.columns([1,1])
    with c1: day=st.date_input("Fecha",value=date.today())
    with c2: division=st.selectbox("División",["TODAS","TW","PM"])
    metrics,rr=day_metrics(day)
    if division=="TODAS":
        st.markdown('<div class="card valle">'+kpi_html(metrics["ALL"])+'</div>',unsafe_allow_html=True)
        for div in ("PM","TW"):
            st.write("")
            base_card(div,metrics[div],[r for r in rr if r["division"]==div])
    else:
        base_card(division,metrics[division],[r for r in rr if r["division"]==division])
    st.warning("El mail renderizado y el PDF ejecutivo se incorporarán en la siguiente iteración cloud.")

elif page=="Empleados":
    st.title("Configuración de empleados")
    emps=employees()
    filter_mode=st.radio("Mostrar",["Solo activos","Todos"],horizontal=True)
    shown=[e for e in emps if filter_mode=="Todos" or e["active"]]
    df=pd.DataFrame([{
        "ID":e["id"],"EMPLEADO":e["full_name"],"ACTIVO":e["active"],
        "CHOFER":e["can_drive"],"AYUDANTE":e["can_assist"],
        "TRELEW":"TW" in e["divisions"],"PUERTO MADRYN":"PM" in e["divisions"]
    } for e in shown])
    edited=st.data_editor(df,hide_index=True,use_container_width=True,num_rows="dynamic",
        disabled=["ID"],column_config={"ID":None})
    c1,c2=st.columns(2)
    if c1.button("Guardar cambios",type="primary",use_container_width=True):
        try:
            original_ids={e["id"] for e in emps}
            for _,r in edited.iterrows():
                payload={"full_name":canon(r["EMPLEADO"]),"active":bool(r["ACTIVO"]),"can_drive":bool(r["CHOFER"]),"can_assist":bool(r["AYUDANTE"])}
                if pd.isna(r.get("ID")):
                    eid=sb().table("employees").insert(payload).execute().data[0]["id"]
                else:
                    eid=r["ID"]; sb().table("employees").update(payload).eq("id",eid).execute()
                    sb().table("employee_divisions").delete().eq("employee_id",eid).execute()
                divs=[]
                if r["TRELEW"]:divs.append("TW")
                if r["PUERTO MADRYN"]:divs.append("PM")
                for d in divs: sb().table("employee_divisions").insert({"employee_id":eid,"division":d}).execute()
            st.success("Base de empleados actualizada."); st.rerun()
        except Exception as e: st.error(str(e))
    del_names=c2.multiselect("Eliminar empleados",options=[e["full_name"] for e in shown])
    if del_names and c2.button("Eliminar seleccionados",use_container_width=True):
        for nm in del_names:
            eid=next(e["id"] for e in emps if e["full_name"]==nm)
            try: sb().table("employees").delete().eq("id",eid).execute()
            except Exception as e: st.error(f"No se pudo eliminar {nm}: {e}")
        st.rerun()

elif page=="Histórico":
    st.title("Histórico de flota y productividad")
    start=st.date_input("Desde",value=date.today().replace(day=1))
    end=st.date_input("Hasta",value=date.today())
    days=sb().table("planning_days").select("*").gte("operational_date",start.isoformat()).lte("operational_date",end.isoformat()).order("operational_date",desc=True).execute().data
    output=[]
    for d in days:
        rr=sb().table("planning_routes").select("*").eq("planning_day_id",d["id"]).execute().data
        for div,cap in (("TW",6),("PM",5)):
            rows=[r for r in rr if r["division"]==div]
            pdv=sum(num(r["pdv"]) for r in rows); b=sum(num(r["packages"]) for r in rows)
            output.append({"FECHA":d["operational_date"],"DIVISIÓN":DIV_NAMES[div],"CAMIONES":len(rows),
                "FLOTA TOTAL":cap,"SIN ASIGNACIÓN":max(cap-len(rows),0),"UTILIZACIÓN %":round(len(rows)/cap*100,1),
                "PDV":pdv,"BULTOS":round(b,1),"DROP SIZE":round(b/pdv,1) if pdv else 0})
    st.dataframe(pd.DataFrame(output),hide_index=True,use_container_width=True)
